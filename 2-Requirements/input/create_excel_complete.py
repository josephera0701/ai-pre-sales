#!/usr/bin/env python3
from openpyxl import Workbook

def create_complete_excel_template():
    wb = Workbook()
    wb.remove(wb.active)
    
    # Sheet 1: Client_Info_Enhanced (Columns A-U from md file)
    ws1 = wb.create_sheet(title='Client_Info')
    ws1.append(['client_id', 'company_name', 'industry_type', 'company_size', 'primary_contact_name', 'primary_contact_email', 'primary_contact_phone', 'technical_contact_name', 'technical_contact_email', 'project_name', 'project_description', 'project_timeline_months', 'budget_range', 'primary_aws_region', 'secondary_aws_regions', 'compliance_requirements', 'business_criticality', 'disaster_recovery_required', 'multi_region_required', 'created_date', 'last_modified'])
    ws1.append(['EXAMPLE-UUID-1234', 'Example Corp', 'Technology', 'Enterprise', 'John Smith', 'john.smith@example.com', '+1-555-0123', 'Jane Doe', 'jane.doe@example.com', 'Cloud Migration Project', 'Migrate legacy applications to AWS', 12, '$100K-$500K', 'us-east-1', 'us-west-2,eu-west-1', 'GDPR,SOC2', 'High', 'Yes', 'Yes', '2024-01-15', '2024-01-15'])
    
    # Sheet 2: Compute_Requirements_Enhanced (Columns A-Z from md file)
    ws2 = wb.create_sheet(title='Compute_Requirements')
    ws2.append(['compute_id', 'client_id', 'server_name', 'environment_type', 'workload_type', 'cpu_cores', 'ram_gb', 'operating_system', 'architecture', 'business_criticality', 'average_utilization_percent', 'peak_utilization_percent', 'scaling_type', 'min_instances', 'max_instances', 'monthly_runtime_hours', 'storage_type', 'root_volume_size_gb', 'additional_storage_gb', 'network_performance', 'placement_group_required', 'dedicated_tenancy_required', 'hibernation_support_required', 'suggested_instance_type', 'estimated_monthly_cost', 'optimization_recommendations'])
    ws2.append(['COMP-UUID-1234', 'EXAMPLE-UUID-1234', 'web-server-01', 'Production', 'Web', 4, 16, 'Amazon Linux', 'x86_64', 'High', 70, 90, 'Auto', 2, 10, 744, 'EBS_GP3', 20, 100, 'High', 'No', 'No', 'No', 'm5.xlarge', '$245.52', 'Consider Reserved Instances'])
    
    # Sheet 3: Storage_Requirements_Enhanced (Columns A-X from md file)
    ws3 = wb.create_sheet(title='Storage_Requirements')
    ws3.append(['storage_id', 'client_id', 'storage_name', 'storage_purpose', 'current_size_gb', 'projected_growth_rate_percent', 'projected_size_12months_gb', 'access_pattern', 'iops_required', 'throughput_mbps_required', 'durability_requirement', 'availability_requirement', 'encryption_required', 'backup_required', 'backup_frequency', 'backup_retention_days', 'cross_region_replication', 'versioning_required', 'lifecycle_management_required', 'compliance_requirements', 'suggested_aws_service', 'suggested_storage_class', 'estimated_monthly_cost', 'optimization_recommendations'])
    ws3.append(['STOR-UUID-1234', 'EXAMPLE-UUID-1234', 'application-data', 'Application_Data', 1000, 20, 1200, 'Frequent', 3000, 250, 'High', 'High', 'Yes', 'Yes', 'Daily', 30, 'No', 'Yes', 'Yes', 'GDPR', 'EBS GP3', 'N/A', '$102.40', 'Consider S3 Intelligent Tiering'])
    
    # Sheet 4: Network_CDN_Enhanced (Columns A-W from md file)
    ws4 = wb.create_sheet(title='Network_CDN')
    ws4.append(['network_id', 'client_id', 'data_transfer_out_gb_monthly', 'data_transfer_in_gb_monthly', 'peak_bandwidth_mbps', 'concurrent_users_expected', 'geographic_distribution', 'load_balancer_count', 'load_balancer_type', 'ssl_certificate_required', 'ssl_certificate_type', 'waf_required', 'ddos_protection_required', 'cdn_required', 'cdn_cache_behavior', 'edge_locations_required', 'api_gateway_required', 'api_calls_monthly', 'vpn_connections_required', 'direct_connect_required', 'bandwidth_direct_connect_mbps', 'estimated_monthly_cost', 'optimization_recommendations'])
    ws4.append(['NET-UUID-1234', 'EXAMPLE-UUID-1234', 500, 200, 1000, 5000, 'us-east-1,eu-west-1', 2, 'ALB', 'Yes', 'Single', 'Yes', 'No', 'Yes', 'Cache_Static', 'us-east-1,eu-west-1', 'Yes', 1000000, 0, 'No', 0, '$425.30', 'Consider CloudFront for global delivery'])
    
    # Sheet 5: Database_Requirements_Enhanced (Columns A-Z from md file)
    ws5 = wb.create_sheet(title='Database_Requirements')
    ws5.append(['database_id', 'client_id', 'database_name', 'database_purpose', 'engine_type', 'engine_version', 'database_size_gb', 'expected_growth_rate_percent', 'instance_class', 'cpu_cores', 'ram_gb', 'storage_type', 'iops_required', 'multi_az_required', 'read_replicas_count', 'read_replica_regions', 'backup_retention_days', 'backup_window_preferred', 'maintenance_window_preferred', 'encryption_at_rest_required', 'encryption_in_transit_required', 'performance_insights_required', 'monitoring_enhanced_required', 'connection_pooling_required', 'estimated_monthly_cost', 'optimization_recommendations'])
    ws5.append(['DB-UUID-1234', 'EXAMPLE-UUID-1234', 'production-db', 'OLTP', 'MySQL', '8.0', 100, 25, 'db.r5.large', 2, 16, 'GP3', 3000, 'Yes', 2, 'us-west-2,eu-west-1', 7, '03:00-04:00', 'Sun:04:00-05:00', 'Yes', 'Yes', 'Yes', 'Yes', 'No', '$485.60', 'Consider Aurora for better performance'])
    
    # Sheet 6: Security_Compliance_Enhanced (Columns A-Z from md file)
    ws6 = wb.create_sheet(title='Security_Compliance')
    ws6.append(['security_id', 'client_id', 'compliance_frameworks', 'data_classification', 'aws_config_required', 'cloudtrail_required', 'cloudtrail_data_events', 'guardduty_required', 'security_hub_required', 'inspector_required', 'macie_required', 'kms_required', 'secrets_manager_required', 'certificate_manager_required', 'iam_access_analyzer_required', 'vpc_flow_logs_required', 'waf_required', 'shield_advanced_required', 'firewall_manager_required', 'network_firewall_required', 'penetration_testing_required', 'vulnerability_scanning_required', 'security_training_required', 'incident_response_plan_required', 'estimated_monthly_cost', 'compliance_gap_analysis'])
    ws6.append(['SEC-UUID-1234', 'EXAMPLE-UUID-1234', 'GDPR,SOC2', 'Confidential', 'Yes', 'Yes', 'Yes', 'Yes', 'Yes', 'Yes', 'No', 'Yes', 'Yes', 'Yes', 'Yes', 'Yes', 'Yes', 'No', 'No', 'No', 'Yes', 'Yes', 'Yes', 'Yes', '$1250.75', 'Implement WAF rules for web protection'])
    
    wb.save('AWS_Cost_Estimation_Template_Enhanced_Complete.xlsx')
    print("Complete Excel file created: AWS_Cost_Estimation_Template_Enhanced_Complete.xlsx")

if __name__ == "__main__":
    create_complete_excel_template()
